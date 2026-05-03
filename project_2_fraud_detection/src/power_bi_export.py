"""
Power BI Export Module - Generate data exports for Power BI integration
"""

import logging
from pathlib import Path
from typing import Dict, Any

import numpy as np
import pandas as pd

import config
from src.utils import setup_logging


class PowerBIExporter:
    """Export data for Power BI integration."""
    
    def __init__(self):
        """Initialize Power BI exporter."""
        self.logger = setup_logging(__name__)
    
    def export_model_metrics(self, results: Dict[str, Dict[str, float]], 
                            output_path: Path = None) -> Path:
        """
        Export model evaluation metrics to Excel.
        
        Args:
            results: Model evaluation results
            output_path: Output file path
            
        Returns:
            Path to exported file
        """
        if output_path is None:
            output_path = config.OUTPUT_DIR / "model_metrics.xlsx"
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        self.logger.info("Exporting model metrics to Excel...")
        
        # Create dataframe
        data = []
        for model_name, metrics in results.items():
            row = {'Model': model_name}
            row.update(metrics)
            data.append(row)
        
        df = pd.DataFrame(data)
        
        # Export to Excel with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Model Metrics', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Model Metrics']
            
            # Format header
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        self.logger.info(f"Model metrics exported to {output_path}")
        return output_path
    
    def export_predictions(self, y_true: np.ndarray, y_pred: np.ndarray, 
                          y_pred_proba: np.ndarray, feature_names: list,
                          output_path: Path = None) -> Path:
        """
        Export predictions with explanations for Power BI.
        
        Args:
            y_true: True labels
            y_pred: Predictions
            y_pred_proba: Prediction probabilities
            feature_names: List of feature names
            output_path: Output file path
            
        Returns:
            Path to exported file
        """
        if output_path is None:
            output_path = config.OUTPUT_DIR / "predictions.xlsx"
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        self.logger.info("Exporting predictions to Excel...")
        
        # Create prediction dataframe
        data = {
            'Index': range(len(y_true)),
            'Actual': y_true,
            'Predicted': y_pred,
            'Is_Correct': (y_true == y_pred).astype(int),
        }
        
        if y_pred_proba is not None:
            if len(y_pred_proba.shape) > 1:
                data['Fraud_Probability'] = y_pred_proba[:, 1]
            else:
                data['Fraud_Probability'] = y_pred_proba
        
        df = pd.DataFrame(data)
        
        # Export to Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Predictions', index=False)
            
            # Format worksheet
            from openpyxl.styles import Font, PatternFill, Alignment
            
            worksheet = writer.sheets['Predictions']
            
            # Header formatting
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Format data cells
            for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
                for i, cell in enumerate(row):
                    if i == 4:  # Is_Correct column
                        if cell.value == 1:
                            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                        else:
                            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    
                    cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        self.logger.info(f"Predictions exported to {output_path}")
        return output_path
    
    def export_feature_statistics(self, X: pd.DataFrame, y: np.ndarray, 
                                 output_path: Path = None) -> Path:
        """
        Export feature statistics by class for Power BI.
        
        Args:
            X: Features dataframe
            y: Target labels
            output_path: Output file path
            
        Returns:
            Path to exported file
        """
        if output_path is None:
            output_path = config.OUTPUT_DIR / "feature_statistics.xlsx"
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        self.logger.info("Exporting feature statistics to Excel...")
        
        # Create statistics dataframe
        stats_data = []
        
        for col in X.select_dtypes(include=[np.number]).columns:
            legitimate = X[y == 0][col]
            fraudulent = X[y == 1][col]
            
            stats_data.append({
                'Feature': col,
                'Legitimate_Mean': legitimate.mean(),
                'Legitimate_Std': legitimate.std(),
                'Legitimate_Min': legitimate.min(),
                'Legitimate_Max': legitimate.max(),
                'Fraudulent_Mean': fraudulent.mean(),
                'Fraudulent_Std': fraudulent.std(),
                'Fraudulent_Min': fraudulent.min(),
                'Fraudulent_Max': fraudulent.max(),
                'Mean_Difference': abs(legitimate.mean() - fraudulent.mean()),
            })
        
        df = pd.DataFrame(stats_data).sort_values('Mean_Difference', ascending=False)
        
        # Export to Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Feature Statistics', index=False)
            
            # Format worksheet
            from openpyxl.styles import Font, PatternFill, Alignment
            
            worksheet = writer.sheets['Feature Statistics']
            
            # Header formatting
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        self.logger.info(f"Feature statistics exported to {output_path}")
        return output_path
    
    def export_confusion_matrix(self, cm: np.ndarray, model_name: str,
                               output_path: Path = None) -> Path:
        """
        Export confusion matrix to Excel.
        
        Args:
            cm: Confusion matrix
            model_name: Model name
            output_path: Output file path
            
        Returns:
            Path to exported file
        """
        if output_path is None:
            output_path = config.OUTPUT_DIR / f"confusion_matrix_{model_name}.xlsx"
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        self.logger.info("Exporting confusion matrix to Excel...")
        
        # Create confusion matrix dataframe
        tn, fp, fn, tp = cm.ravel()
        
        data = {
            'Metric': ['True Negatives', 'False Positives', 'False Negatives', 'True Positives'],
            'Count': [int(tn), int(fp), int(fn), int(tp)],
            'Percentage': [
                tn / (tn + fp) * 100 if (tn + fp) > 0 else 0,
                fp / (tn + fp) * 100 if (tn + fp) > 0 else 0,
                fn / (fn + tp) * 100 if (fn + tp) > 0 else 0,
                tp / (fn + tp) * 100 if (fn + tp) > 0 else 0,
            ]
        }
        
        df = pd.DataFrame(data)
        
        # Export to Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Confusion Matrix', index=False)
            
            # Format worksheet
            from openpyxl.styles import Font, PatternFill, Alignment
            
            worksheet = writer.sheets['Confusion Matrix']
            
            # Header formatting
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Color code cells
            colors = ['#C6EFCE', '#FFC7CE', '#FFC7CE', '#C6EFCE']
            for i, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df) + 1), 0):
                row[2].fill = PatternFill(start_color=colors[i], end_color=colors[i], fill_type='solid')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        self.logger.info(f"Confusion matrix exported to {output_path}")
        return output_path
    
    def export_all(self, results: Dict[str, Dict[str, float]], 
                  y_true: np.ndarray, y_pred: np.ndarray,
                  y_pred_proba: np.ndarray, X: pd.DataFrame,
                  feature_names: list, cm: np.ndarray,
                  model_name: str = "best_model") -> Dict[str, Path]:
        """
        Export all data for Power BI.
        
        Args:
            results: Model evaluation results
            y_true: True labels
            y_pred: Predictions
            y_pred_proba: Prediction probabilities
            X: Features dataframe
            feature_names: List of feature names
            cm: Confusion matrix
            model_name: Model name
            
        Returns:
            Dictionary of exported file paths
        """
        self.logger.info("Exporting all data for Power BI...")
        
        exported_files = {
            'model_metrics': self.export_model_metrics(results),
            'predictions': self.export_predictions(y_true, y_pred, y_pred_proba, feature_names),
            'feature_statistics': self.export_feature_statistics(X, y_true),
            'confusion_matrix': self.export_confusion_matrix(cm, model_name),
        }
        
        self.logger.info("All data exported successfully!")
        return exported_files


if __name__ == "__main__":
    logger = setup_logging(__name__)
    logger.info("Power BI Export module loaded successfully")
